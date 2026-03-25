import os
from openpyxl.styles import PatternFill

# ================================
# Path Configuration
# ================================
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
AUDIT_DIR  = os.path.join(BASE_DIR, 'audit_files')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ================================
# File Names
# — Change these to match your audit file names
# ================================
PREVIOUS_AUDIT = os.path.join(AUDIT_DIR, 'PIIAudit_110326.xlsx')
CURRENT_AUDIT  = os.path.join(AUDIT_DIR, 'PIIAudit_170326.xlsx')

# ================================
# Column Names
# — Change these if your audit headers differ
# — Or leave as is — fuzzy matching will auto detect
# ================================
COL_DATABASE = 'Database'
COL_SCHEMA   = 'Schema'
COL_TABLE    = 'Tables'
COL_COLUMN   = 'Columns'
COL_DATATYPE = 'Datatype'
COL_ISPII    = 'IsPII'

# ================================
# IsPII Normalisation Values
# — Add more if your audit uses different values
# ================================
YES_VALUES = ['yes', 'y', 'true', '1', 'pii']
NO_VALUES  = ['no',  'n', 'false', '0', 'non-pii', 'nonpii']

# ================================
# Change Types
# ================================
NEW_DATABASE     = 'New Database Added'
DROPPED_DATABASE = 'Database Dropped'
NEW_SCHEMA       = 'New Schema Added'
DROPPED_SCHEMA   = 'Schema Dropped'
NEW_TABLE        = 'New Table Added'
DROPPED_TABLE    = 'Table Dropped'
NEW_COLUMN       = 'New Column Added'
DROPPED_COLUMN   = 'Column Dropped'
DATATYPE_CHANGED = 'Datatype Changed'

# ================================
# Severity Levels
# ================================
HIGH   = 'HIGH'
MEDIUM = 'MEDIUM'
LOW    = 'LOW'

# ================================
# Highlight Colours — ARGB format
# ================================
RED    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
ORANGE = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
BLUE   = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")
PURPLE = PatternFill(start_color="FFDDA0DD", end_color="FFDDA0DD", fill_type="solid")

# ================================
# Colour Map
# ================================
COLOUR_MAP = {
    NEW_COLUMN       : {"Yes": RED,    "No": YELLOW},
    NEW_TABLE        : {"Yes": RED,    "No": YELLOW},
    NEW_SCHEMA       : {"Yes": RED,    "No": YELLOW,   "Unknown": RED},
    NEW_DATABASE     : {"Yes": RED,    "No": YELLOW,   "Unknown": RED},
    DROPPED_COLUMN   : {"Yes": ORANGE, "No": BLUE},
    DROPPED_TABLE    : {"Yes": ORANGE, "No": BLUE,     "N/A": BLUE},
    DROPPED_SCHEMA   : {"Yes": ORANGE, "No": BLUE,     "N/A": BLUE},
    DROPPED_DATABASE : {"Yes": ORANGE, "No": BLUE,     "N/A": BLUE},
    DATATYPE_CHANGED : {"Yes": RED,    "No": PURPLE},
}
