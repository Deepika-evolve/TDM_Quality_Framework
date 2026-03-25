import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from config import COLOUR_MAP, HIGH, COL_DATABASE, COL_SCHEMA, COL_TABLE, COL_COLUMN, COL_DATATYPE, COL_ISPII
from metrics import get_drift_metrics
from data_quality import detect_headers, clean_dataframe
from utils import get_logger
logger = get_logger("Generating Drift Detection File")
COL_KEY_MAP = {
    COL_DATABASE : 'Database',
    COL_SCHEMA   : 'Schema',
    COL_TABLE    : 'Table',
    COL_COLUMN   : 'Column',
    COL_DATATYPE : 'Datatype',
    COL_ISPII    : 'IsPII',
}


def write_excel_report(results, current_file, output_file):

    if not results:
        ##print("\nNo changes — output file not generated")
        logger.info("\nNo changes — output file not generated")
        return

    xl = pd.ExcelFile(current_file)
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet in xl.sheet_names:
            df = pd.read_excel(current_file, sheet_name=sheet)
            # Dedup before writing — remove duplicates from output
            df = clean_dataframe(df, detect_headers(df), mode='dedup')
            df.to_excel(writer, sheet_name=sheet, index=False)
    try:
        wb = load_workbook(output_file)

        for sheet_name in wb.sheetnames:
            ws      = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]

            for col in ['ChangeType', 'Severity', 'Action']:
                if col not in headers:
                    ws.cell(row=1, column=len(headers) + 1).value = col
                    headers.append(col)

            change_col   = headers.index('ChangeType') + 1
            severity_col = headers.index('Severity')   + 1
            action_col   = headers.index('Action')     + 1

            col_positions = {}
            for col_name in [COL_DATABASE, COL_SCHEMA, COL_TABLE, COL_COLUMN, COL_DATATYPE, COL_ISPII]:
                if col_name in headers:
                    col_positions[col_name] = headers.index(col_name) + 1

            sheet_results = [r for r in results if r.get('Sheet') == sheet_name]
            sheet_results = sorted(sheet_results, key=lambda x: x['ChangeType'])

            for result in sheet_results:
                row_number  = result.get('ExcelRowNumber')
                change_type = result['ChangeType']
                is_pii      = str(result['IsPII'])

                if row_number is None:
                    continue

                colour = COLOUR_MAP.get(change_type, {}).get(is_pii)
                if colour:
                    for cell in ws[row_number]:
                        cell.fill = colour

                ws.cell(row=row_number, column=change_col).value   = result['ChangeType']
                ws.cell(row=row_number, column=severity_col).value = result['Severity']
                ws.cell(row=row_number, column=action_col).value   = result['Action']

            dropped = [r for r in sheet_results if 'Dropped' in r['ChangeType']]

            if dropped:
                sep_row = ws.max_row + 1
                ws.cell(row=sep_row, column=1).value = '=== DROPPED ==='
                ws.cell(row=sep_row, column=1).font  = Font(bold=True)

                for result in dropped:
                    last_row = ws.max_row + 1

                    for col_name, position in col_positions.items():
                        result_key = COL_KEY_MAP.get(col_name, col_name)
                        value      = result.get(result_key, 'N/A')
                        ws.cell(last_row, position).value = value

                    ws.cell(last_row, change_col).value   = result['ChangeType']
                    ws.cell(last_row, severity_col).value = result['Severity']
                    ws.cell(last_row, action_col).value   = result['Action']

                    colour = COLOUR_MAP.get(result['ChangeType'], {}).get(str(result['IsPII']))
                    if colour:
                        for col in range(1, len(headers) + 1):
                            ws.cell(last_row, col).fill = colour

        _write_drift_summary(wb, results)
        _write_critical_changes(wb, results)
        drift_idx= wb.sheetnames.index('Consolidated_Drift_Summary')
        wb.move_sheet('Consolidated_Drift_Summary',offset=-drift_idx)
        critical_idx = wb.sheetnames.index('Consolidated_Critical_Changes')
        wb.move_sheet('Consolidated_Critical_Changes', offset=-(critical_idx-1))
        wb.save(output_file)
        ##_print_summary(results)
        ##print(f"\nExcel report saved — {output_file}")
        logger.info(f"\nExcel report saved — {output_file}")
    except PermissionError:
            logger.error(f"Cannot save Excel report - file is open: {output_file}\nPlease close the file and retry.")
    except FileNotFoundError as e:
            logger.error(f"Output file path not found: {e}")
    except Exception as e:
            logger.error(f"Error generating Excel report: {e}")

def _write_drift_summary(wb, results):
    metrics = get_drift_metrics(results)

    if 'Consolidated_Drift_Summary' in wb.sheetnames:
        del wb['Consolidated_Drift_Summary']

    ws     = wb.create_sheet('Consolidated_Drift_Summary')
    GREY   = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    HEADER = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")

    ws.append(['Metric', 'Count'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER

    sections = {
        'DATABASE LEVEL' : ['New Databases', 'Dropped Databases'],
        'SCHEMA LEVEL'   : ['New Schemas', 'Dropped Schemas'],
        'TABLE LEVEL'    : ['New Tables', 'Dropped Tables'],
        'COLUMN LEVEL'   : ['New PII Columns', 'New Non PII Columns',
                            'Dropped PII Columns', 'Dropped Non PII Columns',
                            'Datatype Changes'],
        'SEVERITY'       : ['HIGH Severity', 'MEDIUM Severity', 'LOW Severity'],
    }

    for section, keys in sections.items():
        ws.append([section, ''])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = GREY
        for key in keys:
            ws.append([key, metrics.get(key, 0)])


def _write_critical_changes(wb, results):
    results_df  = pd.DataFrame(results)
    critical_df = results_df[results_df['Severity'] == HIGH].drop(columns=['ExcelRowNumber'])

    if critical_df.empty:
        return

    if 'Consolidated_Critical_Changes' in wb.sheetnames:
        del wb['Consolidated_Critical_Changes']

    ws      = wb.create_sheet('Consolidated_Critical_Changes')
    headers = list(critical_df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ##RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    for _, row in critical_df.iterrows():
        ws.append(list(row))
        ##for cell in ws[ws.max_row]:
            ##cell.fill = RED


def _print_summary(results):
    metrics  = get_drift_metrics(results)
    sections = {
        'DATABASE LEVEL' : ['New Databases', 'Dropped Databases'],
        'SCHEMA LEVEL'   : ['New Schemas', 'Dropped Schemas'],
        'TABLE LEVEL'    : ['New Tables', 'Dropped Tables'],
        'COLUMN LEVEL'   : ['New PII Columns', 'New Non PII Columns',
                            'Dropped PII Columns', 'Dropped Non PII Columns',
                            'Datatype Changes'],
        'SEVERITY'       : ['HIGH Severity', 'MEDIUM Severity', 'LOW Severity'],
    }
    print(f"\n{'='*45}")
    print(f"  DRIFT SUMMARY")
    print(f"{'='*45}")
    for section, keys in sections.items():
        print(f"\n  {section}")
        for key in keys:
            print(f"    {key:25}: {metrics.get(key, 0)}")
    print(f"\n{'='*45}")
