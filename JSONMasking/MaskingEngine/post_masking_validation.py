import json
import logging
import pandas as pd
from openpyxl.styles import PatternFill, Font
from PIIAnalysisEngine.pii_analyser import analyse_json
from QualityEngine import validate_json_string, partial_mask_value,clean_value
import jm_config as config
import math
logger = logging.getLogger(__name__)


def extract_masked_values(masked_df, classifier, json_col):
    """Traverse masked JSON — return dict of doc_id|path → (value, type)"""
    masked_values = {}
    for idx, row in masked_df.iterrows():
        json_str = str(row[json_col])
        doc_id   = f'row_{idx+1}'
        try:
            if json_str.strip().startswith('['):
                doc_list = json.loads(json_str)
                for i, item in enumerate(doc_list):
                    if isinstance(item, dict):
                        for finding in analyse_json(item, classifier, f'[{i}]'):
                            key = f'{doc_id}|{finding["path"]}'
                            masked_values[key] = (finding['value'],
                                                  type(finding['value']).__name__)
            else:
                doc = validate_json_string(json_str, doc_id)
                if doc is None: continue
                for finding in analyse_json(doc, classifier):
                    key = f'{doc_id}|{finding["path"]}'
                    masked_values[key] = (finding['value'],
                                          type(finding['value']).__name__)
        except Exception as e:
            logger.warning(f'Error traversing masked row {idx}: {e}')
    return masked_values


def generate_masking_report(hash_audit_df, masked_df, classifier,
                            report_path, json_col):
    """Generate strong validation report — type check + masked check + metrics"""
    logger.info('Generating masking validation report...')

    masked_values = extract_masked_values(masked_df, classifier, json_col)

    approved = hash_audit_df[
        (hash_audit_df['is_pii'].str.upper() == 'YES') &
        (hash_audit_df['approved'].str.upper() == 'YES') &
        (hash_audit_df['is_pii'] != 'container')
    ]

    rows = []
    for _, row in approved.iterrows():
        try:
            parts = str(row['doc_id']).split('_')
            lookup_doc_id = f'{parts[0]}_{parts[1]}'
            lookup_key    = f'{lookup_doc_id}|{row["path"]}'

            original_raw = row['value']

            # Check if original value is empty/null/NaN
            is_empty = (
                    original_raw is None or
                    (isinstance(original_raw, float) and math.isnan(original_raw)) or
                    str(original_raw).strip() == ''
            )

            if is_empty:
                # Empty field — no PII to validate — skip
                original_disp = ''
                original_type = 'empty'
                masked_val = ''
                masked_type = 'empty'
                mask_status = 'Skip'
                type_status = 'Skip'
                expected_mask = 'Skip'
                actual_mask = 'Skip'
            else:
                original_disp = partial_mask_value(original_raw,
                                                   show_pii=not config.PARTIAL_MASK_IN_AUDIT)
                original_type = type(original_raw).__name__

                masked_result = masked_values.get(lookup_key)
                if masked_result:
                    masked_val = masked_result[0]
                    masked_type = masked_result[1]
                else:
                    masked_val = 'not found'
                    masked_type = 'unknown'

                expected_mask = 'Yes'
                actual_mask = 'Yes' if (str(masked_val) != str(original_raw)
                                        and masked_val != 'not found') else 'No'
                mask_status = 'Pass' if expected_mask == actual_mask else 'Fail'
                type_status = 'Pass' if original_type == masked_type else 'Fail'
            if mask_status == 'Fail' and str(masked_val) == str(original_raw):
                # Value unchanged — could be wrong pattern or empty
                mask_status = 'Invalid input data format'

            rows.append({
                'doc_id'         : row['doc_id'],
                'path'           : row['path'],
                'key'            : row['key'],
                'original_value' : original_disp,
                'masked_value'   : masked_val,
                'original_type'  : original_type,
                'masked_type'    : masked_type,
                'type_status'    : type_status,
                'expected_mask'  : expected_mask,
                'actual_mask'    : actual_mask,
                'mask_status'    : mask_status,
                'is_pii'         : row['is_pii'],
                'approved'       : row['approved'],
                'approved_by'    : row.get('approved_by', ''),
                'mask_function'  : row.get('mask_function', '')
            })
        except Exception as e:
            logger.warning(f'Validation skipped for path {row["path"]}: {e}')
            continue

    report_df = pd.DataFrame(rows)
    if 'source_file' in masked_df.columns and 'doc_id' in masked_df.columns:
        source_lookup = masked_df[['doc_id', 'source_file']].drop_duplicates()
        report_df[['docid', 'id', 'rownum']] = report_df['doc_id'].str.split('_', expand=True)
        report_df['lookup_doc_id'] = report_df['docid'].astype(str) + '_' + report_df['id'].astype(str)
        report_df = report_df.merge(source_lookup,  left_on='lookup_doc_id',  # Column name in report_df
                                    right_on='doc_id',      # Column name in source_lookup
                                    how='left')
        report_df['source_file'] = report_df['source_file'].fillna('excel_input')
        report_df.drop(columns=['docid', 'id', 'rownum', 'lookup_doc_id', 'doc_id_y'], inplace=True)
        report_df.rename(columns={'doc_id_x': 'doc_id'}, inplace=True)
    else:
        report_df['source_file'] = 'excel_input'
    total = len(report_df)
    skipped = len(report_df[report_df['mask_status'] == 'Skip'])
    wrong_pattern = len(report_df[report_df['mask_status'] == 'Invalid input data format'])
    validated = total - skipped - wrong_pattern
    mask_pass = len(report_df[report_df['mask_status'] == 'Pass'])
    mask_fail = len(report_df[report_df['mask_status'] == 'Fail'])
    type_pass = len(report_df[report_df['type_status'] == 'Pass'])
    type_fail = len(report_df[report_df['type_status'] == 'Fail'])
    pass_rate = round((mask_pass / validated * 100), 1) if validated > 0 else 0

    summary_df = pd.DataFrame([{
        'Total PII Fields': total,
        'Validated': validated,
        'Skipped (empty)': skipped,
        'Invalid input data format': wrong_pattern,
        'Masking Pass': mask_pass,
        'Masking Fail': mask_fail,
        'Masking Pass Rate %': f'{pass_rate}%',
        'Type Match Pass': type_pass,
        'Type Match Fail': type_fail,
    }])

    # Write summary then detail to Excel
    with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        report_df.to_excel(writer, sheet_name='Details', index=False)

    wb = writer.book

    # Colour code Details sheet
    ws_detail = wb['Details']
    green_fill = PatternFill(start_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', fill_type='solid')
    orange_fill = PatternFill(start_color='FFEB9C', fill_type='solid')

    # Find column indices for mask_status and type_status
    header = [ws_detail.cell(1, c).value for c in
              range(1, ws_detail.max_column + 1)]
    mask_col = header.index('mask_status') + 1 if 'mask_status' in header else None
    type_col = header.index('type_status') + 1 if 'type_status' in header else None

    for row in ws_detail.iter_rows(min_row=2):
        if mask_col:
            cell = ws_detail.cell(row[0].row, mask_col)
            if cell.value == 'Pass':
                cell.fill = green_fill
            elif cell.value == 'Fail':
                cell.fill = red_fill
        if type_col:
            cell = ws_detail.cell(row[0].row, type_col)
            if cell.value == 'Pass':
                cell.fill = green_fill
            elif cell.value == 'Fail':
                cell.fill = red_fill

    # Colour code Summary sheet
    ws_sum = wb['Summary']
    for row in ws_sum.iter_rows(min_row=2):
        for cell in row:
            col_name = ws_sum.cell(1, cell.column).value
            if col_name == 'Masking Fail' or col_name == 'Type Match Fail':
                if cell.value and cell.value > 0:
                    cell.fill = red_fill
                else:
                    cell.fill = green_fill
            elif col_name == 'Masking Pass Rate %':
                val = float(str(cell.value).replace('%', ''))
                cell.fill = green_fill if val >= 95 else red_fill
            elif col_name == 'Wrong Pattern':
                if cell.value and cell.value > 0:
                    cell.fill = orange_fill  # orange — input data quality issue
                else:
                    cell.fill = green_fill

            # Add orange fill definition at top
            orange_fill = PatternFill(start_color='FFEB9C', fill_type='solid')

    logger.info(f'Masking validation Report: {report_path}')
    logger.info(f'Total: {total} | Pass: {mask_pass} | Skipped (empty): {skipped}|Invalid input data format: {wrong_pattern} | Fail: {mask_fail} | Rate: {pass_rate}%')
    return report_df
